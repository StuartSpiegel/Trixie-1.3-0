from jira.client import JIRA
from openpyxl import Workbook

'''The purpose of this script is to identify issues with parsing fields into JIRA including custom fields, 
a report using the workbook library is generated upon execution 
'''

# Global variables
issues = []

# Basic tuple sign in
jira_options = {'server': 'http://example.com'}


def cookie_auth():
    auth_jira = JIRA(auth=('username', 'password'))


def OAuth():
    key_cert_data = None
    with open(key_cert_data, 'r') as key_cert_file:
        key_cert_data = key_cert_file.read()

    oauth_dict = {
        'access_token': 'foo',
        'access_token_secret': 'bar',
        'consumer_key': 'jira-oauth-consumer',
        'key_cert': key_cert_data
    }

    auth_jira = JIRA(oauth=oauth_dict)


# Create the initial jira object
jira = JIRA(options=jira_options, basic_auth=('admin', 'password'))
props = jira.application_properties()  # requires JIRA system administrator permissions

# all issues reported by the admin user
issues = jira.search_issues("assignee=admin")

key_list = []
summary_list = []
story_list = []
story_category_list = []
description_list = []
acceptance_criteria_list = []
testing_list = []
considerations_list = []

project_issues = jira.search_issues('project=ALPHA AND assignee= admin')

for issue in project_issues:
    key_list.append(issue.key)
    summary_list.append(issue.fields.summary)
    description_list.append(issue.fields)
    acceptance_criteria_list.append(issue.fields)
    considerations_list.append(issue.fields)
    testing_list.append(issue.fields)
    story_category_list.append(issue.fields)
    story_list.append(issue.fields)

# Set the workbook variables for report generation
wb = Workbook()
ws = wb.active()
key_row = 1
summary_row = 1
description_row = 1
considerations_row = 1
acceptance_criteria_row = 1
testing_row = 1
story_list_row = 1
story_category_row = 1

start_column = 1

for key in key_list:
    ws.cell(row=key_row, column=start_column).value = key
    key_row += 1

# add fields here
for description in description_list:
    ws.cell(row=description_row, column=start_column + 2).value = description
    description_row += 1

wb.save("JIRA-REPORT")
